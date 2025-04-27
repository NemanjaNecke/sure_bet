import argparse
import glob
import os
import shutil
import pickle
import pandas as pd

# RAPIDFUZZ (faster alternative to fuzzywuzzy)
# pip install rapidfuzz
from rapidfuzz import process, fuzz

import openpyxl
import logging
import datetime
import zipfile
from typing import Dict, List, Tuple, Optional

# For partial concurrency in data loading
from concurrent.futures import ThreadPoolExecutor, as_completed

# If you still prefer or rely on fuzzywuzzy, you can swap these lines:
# from fuzzywuzzy import process, fuzz

# ---------------------------------------------------- #
#                  Global Configuration               #
# ---------------------------------------------------- #
LOG_DIR = "log"
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    filename=os.path.join(LOG_DIR, "sure_bet_strict.log"),
    filemode="w",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# Example config defaults (override with argparse or .yaml if desired)
PICKLE_FOLDER_PATH = "./pickle_data"   # Folder containing .pkl
EXCEL_FOLDER_PATH = "./excel_data"     # Folder containing .xlsx
FUZZY_THRESHOLD = 80
TOTAL_STAKE = 100
OUTPUT_FOLDER = "./results"
ARCHIVE_FOLDER = "./archive"

# Ensure output and archive directories exist
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(ARCHIVE_FOLDER, exist_ok=True)

# Required columns for DataFrame
REQUIRED_COLUMNS = ['home', '1', 'x', '2', 'time']

# ---------------------------------------------------- #
#         Concurrency-Enhanced File Loading           #
# ---------------------------------------------------- #
def load_pickle_file(file_path: str) -> Tuple[str, pd.DataFrame]:
    """Load a single .pkl file and return (bookie_name, DataFrame)."""
    file_name = os.path.basename(file_path)
    bookie_name = os.path.splitext(file_name)[0]
    try:
        with open(file_path, "rb") as f:
            df = pickle.load(f)
        logging.info(f"Loaded pickle file: {file_name}")
        return bookie_name, df
    except Exception as e:
        logging.error(f"Could not load {file_path}. Error: {e}")
        return bookie_name, pd.DataFrame()  # Return empty on error

def load_excel_file(file_path: str) -> Tuple[str, pd.DataFrame]:
    """Load a single .xlsx file and return (bookie_name, DataFrame)."""
    file_name = os.path.basename(file_path)
    bookie_name = os.path.splitext(file_name)[0]
    try:
        df = pd.read_excel(file_path, engine="openpyxl")
        logging.info(f"Loaded Excel file: {file_name}")
        return bookie_name, df
    except Exception as e:
        logging.error(f"Could not load {file_path}. Error: {e}")
        return bookie_name, pd.DataFrame()

def load_pickle_data_concurrent(folder_path: str) -> Dict[str, pd.DataFrame]:
    """Load all .pkl files in concurrent threads and return {bookie_name: df}."""
    bookies = {}
    pkl_files = glob.glob(os.path.join(folder_path, "*.pkl"))

    if not pkl_files:
        logging.warning(f"No .pkl files found in {folder_path}")
        return bookies

    with ThreadPoolExecutor() as executor:
        future_to_file = {executor.submit(load_pickle_file, f): f for f in pkl_files}
        for future in as_completed(future_to_file):
            bookie_name, df = future.result()
            if not df.empty:
                bookies[bookie_name] = df

    return bookies

def load_excel_data_concurrent(folder_path: str) -> Dict[str, pd.DataFrame]:
    """Load all .xlsx files in concurrent threads and return {bookie_name: df}."""
    bookies = {}
    xlsx_files = (
        glob.glob(os.path.join(folder_path, "*.xlsx")) +
        glob.glob(os.path.join(folder_path, "*.XLSX"))
    )

    if not xlsx_files:
        logging.warning(f"No .xlsx files found in {folder_path}")
        return bookies

    with ThreadPoolExecutor() as executor:
        future_to_file = {executor.submit(load_excel_file, f): f for f in xlsx_files}
        for future in as_completed(future_to_file):
            bookie_name, df = future.result()
            if not df.empty:
                bookies[bookie_name] = df

    return bookies

# ---------------------------------------------------- #
#            DataFrame Standardization                #
# ---------------------------------------------------- #
def standardize_columns(df: pd.DataFrame, bookie_name: str) -> pd.DataFrame:
    """
    Standardizes column names. If you have known mappings per bookie, do them here.
    Otherwise, do a generic best-effort approach.
    """
    # Example of known mapping (expand as needed)
    known_mappings = {
        'sportplus': {
            'kickoff': 'time',
            '1.00': '1',
            '2.00': '2',
        },
        # Add more if needed
    }

    # If a known mapping is found, apply it
    applied_mapping = {}
    for key, mapping in known_mappings.items():
        if key.lower() in bookie_name.lower():
            applied_mapping = mapping
            break

    # Renaming columns
    if applied_mapping:
        df = df.rename(columns=applied_mapping)

    # Force column names to lower for easier matching
    df.columns = df.columns.str.lower()

    # Attempt to rename to canonical column names
    rename_dict = {}
    for col in df.columns:
        if col in ['1', '1.0', '1.00']:
            rename_dict[col] = '1'
        elif col in ['x', 'x.0']:
            rename_dict[col] = 'x'
        elif col in ['2', '2.0', '2.00']:
            rename_dict[col] = '2'
        elif col in ['home team', 'home_team']:
            rename_dict[col] = 'home'
        elif col in ['time', 'kickoff']:
            rename_dict[col] = 'time'
    df = df.rename(columns=rename_dict)

    # Remove duplicates
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated()]

    return df

def extract_relevant_columns(df: pd.DataFrame, bookie_name: str) -> pd.DataFrame:
    """Keep only relevant columns, ensure numeric odds, remove duplicates, etc."""
    df = standardize_columns(df, bookie_name)

    # Check required columns
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"{bookie_name} is missing columns: {missing}")

    # Convert odds to numeric
    for col in ['1', 'x', '2']:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.'), errors='coerce')

    # Convert time to datetime
    df['time'] = pd.to_datetime(df['time'], errors='coerce')

    # Keep only required columns
    df = df[REQUIRED_COLUMNS].copy()

    # Remove duplicates by sum of odds
    df['sum_odds'] = df['1'] + df['x'] + df['2']
    df.sort_values(by='sum_odds', inplace=True)
    df = df.groupby(['home', 'time'], as_index=False).first()
    df.drop(columns=['sum_odds'], inplace=True, errors='ignore')

    # Drop invalid or NaN rows
    initial_count = len(df)
    df.dropna(subset=['home', 'time', '1', 'x', '2'], inplace=True)
    df = df[(df['1'] > 0) & (df['x'] > 0) & (df['2'] > 0)]
    final_count = len(df)
    if final_count < initial_count:
        logging.info(f"{bookie_name}: Dropped {initial_count - final_count} rows with invalid data.")

    df.reset_index(drop=True, inplace=True)
    return df

# ---------------------------------------------------- #
#            Fuzzy Merging / Multi-way Merge          #
# ---------------------------------------------------- #
def fuzzy_merge_two_dfs(
    df_master: pd.DataFrame,
    df_new: pd.DataFrame,
    new_bookie_name: str,
    threshold: int = 80
) -> pd.DataFrame:
    """Fuzzy-merge df_new (home,time,1,x,2) into df_master, adding {bookie}_1, etc."""
    if df_new.empty:
        # Just add empty columns
        for col_suffix in ['_1', '_x', '_2']:
            df_master[f"{new_bookie_name}{col_suffix}"] = None
        return df_master

    # Create a single string key for each row (home + time) in both DFs
    df_master['home_time'] = df_master.apply(lambda r: f"{r['home']} {r['time']}", axis=1)
    df_new['home_time'] = df_new.apply(lambda r: f"{r['home']} {r['time']}", axis=1)

    # Convert df_new to a dict for quick lookups
    df_new_dict = df_new.set_index('home_time').to_dict('index')
    df_new_list = list(df_new_dict.keys())

    # Optionally, handle exact matches first to skip fuzzy for those
    # This can drastically reduce calls to fuzzy matching.
    exact_matches_master = set(df_master['home_time']).intersection(df_new_list)
    # Mark them so we skip fuzzy for exact matches
    df_master['exact_match'] = df_master['home_time'].isin(exact_matches_master)

    merged_rows = []
    total_rows = len(df_master)
    fuzzy_count = 0
    matched_count = 0

    for _, row in df_master.iterrows():
        home_time_master = row['home_time']
        out_row = row.to_dict()
        if row['exact_match']:
            # Direct exact match
            row_n = df_new_dict[home_time_master]
            out_row[f"{new_bookie_name}_1"] = row_n['1']
            out_row[f"{new_bookie_name}_x"] = row_n['x']
            out_row[f"{new_bookie_name}_2"] = row_n['2']
            merged_rows.append(out_row)
            matched_count += 1
        else:
            # Fuzzy match
            fuzzy_count += 1
            best_match, best_score = process.extractOne(
                home_time_master,
                df_new_list,
                scorer=fuzz.token_set_ratio
            ) or (None, None)
            if best_match and best_score >= threshold:
                row_n = df_new_dict[best_match]
                out_row[f"{new_bookie_name}_1"] = row_n['1']
                out_row[f"{new_bookie_name}_x"] = row_n['x']
                out_row[f"{new_bookie_name}_2"] = row_n['2']
                matched_count += 1
            else:
                out_row[f"{new_bookie_name}_1"] = None
                out_row[f"{new_bookie_name}_x"] = None
                out_row[f"{new_bookie_name}_2"] = None
            merged_rows.append(out_row)

    df_merged = pd.DataFrame(merged_rows)
    df_merged.drop(columns=['home_time', 'exact_match'], inplace=True, errors='ignore')

    logging.info(
        f"[Merge {new_bookie_name}] Total rows in master: {total_rows}, "
        f"Exact matches: {total_rows - fuzzy_count}, "
        f"Fuzzy attempts: {fuzzy_count}, "
        f"Successfully matched: {matched_count}"
    )
    return df_merged

def multiway_fuzzy_merge(bookies: Dict[str, pd.DataFrame], threshold: int = 80) -> pd.DataFrame:
    """Iteratively fuzzy-merge all bookies into one wide DataFrame."""
    if not bookies:
        return pd.DataFrame()

    # Convert dict to list
    items = list(bookies.items())

    # Start with the first as master
    base_name, df_master = items[0]
    # Rename columns from (1,x,2) -> (base_1, base_x, base_2)
    df_master = df_master.rename(
        columns={
            '1': f"{base_name}_1",
            'x': f"{base_name}_x",
            '2': f"{base_name}_2",
        }
    )

    # Master must still have 'home' and 'time'
    if 'home' not in df_master.columns or 'time' not in df_master.columns:
        logging.error(f"{base_name} missing 'home' or 'time'. Cannot proceed.")
        return pd.DataFrame()

    for bookie_name, df_new in items[1:]:
        required = {'home', 'time', '1', 'x', '2'}
        if not required.issubset(df_new.columns):
            logging.warning(f"{bookie_name} is missing required columns; skipping.")
            continue
        df_master = fuzzy_merge_two_dfs(
            df_master=df_master,
            df_new=df_new,
            new_bookie_name=bookie_name,
            threshold=threshold
        )

    return df_master

# ---------------------------------------------------- #
#    Sure Bets Calculation / Staking Formula          #
# ---------------------------------------------------- #
from sympy import symbols, Eq, solve

def beat_bookies(odds1, odds2, odds3, total_stake=100):
    """Allocate total_stake across 1/x/2 so all outcomes yield same profit."""
    try:
        x, y, z = symbols('x y z', real=True, nonnegative=True)
        eq1 = Eq(x + y + z, total_stake)
        eq2 = Eq(odds1*x, odds2*y)
        eq3 = Eq(odds1*x, odds3*z)
        sol = solve((eq1, eq2, eq3), (x, y, z), dict=True)
        if not sol:
            return { 'Stake1': 0, 'Stake2': 0, 'Stake3': 0, 'Profit1': 0, 'Profit2': 0, 'Profit3': 0 }
        s = sol[0]
        stake1, stake2, stake3 = s[x], s[y], s[z]
        if stake1 < 0 or stake2 < 0 or stake3 < 0:
            return { 'Stake1': 0, 'Stake2': 0, 'Stake3': 0, 'Profit1': 0, 'Profit2': 0, 'Profit3': 0 }
        return {
            'Stake1': float(stake1),
            'Stake2': float(stake2),
            'Stake3': float(stake3),
            'Profit1': float(odds1 * stake1 - total_stake),
            'Profit2': float(odds2 * stake2 - total_stake),
            'Profit3': float(odds3 * stake3 - total_stake)
        }
    except Exception as e:
        logging.error(f"Error in beat_bookies: {e}")
        return { 'Stake1': 0, 'Stake2': 0, 'Stake3': 0, 'Profit1': 0, 'Profit2': 0, 'Profit3': 0 }

def calculate_implied_probabilities(df_merged: pd.DataFrame, total_stake: float) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Compute best odds, implied probability, and extract sure bets."""
    if df_merged.empty:
        return df_merged, pd.DataFrame()

    all_cols = df_merged.columns
    odds_1_cols = [c for c in all_cols if c.endswith('_1')]
    odds_x_cols = [c for c in all_cols if c.endswith('_x')]
    odds_2_cols = [c for c in all_cols if c.endswith('_2')]

    if not odds_1_cols or not odds_x_cols or not odds_2_cols:
        logging.error("Missing or invalid odds columns; cannot calculate sure bets.")
        return df_merged, pd.DataFrame()

    # Find best (lowest) odds for each outcome
    df_merged['best_1'] = df_merged[odds_1_cols].min(axis=1, skipna=True)
    df_merged['best_x'] = df_merged[odds_x_cols].min(axis=1, skipna=True)
    df_merged['best_2'] = df_merged[odds_2_cols].min(axis=1, skipna=True)

    # Identify which bookie had the best odds
    try:
        df_merged['best_1_source'] = df_merged[odds_1_cols].idxmin(axis=1).str.replace('_1','')
        df_merged['best_x_source'] = df_merged[odds_x_cols].idxmin(axis=1).str.replace('_x','')
        df_merged['best_2_source'] = df_merged[odds_2_cols].idxmin(axis=1).str.replace('_2','')
    except Exception as e:
        logging.warning(f"Error identifying best odds source: {e}")
        df_merged['best_1_source'] = None
        df_merged['best_x_source'] = None
        df_merged['best_2_source'] = None

    # Drop rows with invalid best odds
    valid_mask = (df_merged['best_1'] > 0) & (df_merged['best_x'] > 0) & (df_merged['best_2'] > 0)
    df_valid = df_merged[valid_mask].copy()
    excluded = len(df_merged) - len(df_valid)
    if excluded > 0:
        logging.info(f"Excluded {excluded} rows with missing or zero best odds.")

    # Calculate implied
    df_valid['implied'] = (1/df_valid['best_1']) + (1/df_valid['best_x']) + (1/df_valid['best_2'])

    # Filter sure bets
    sure_bets = df_valid[df_valid['implied'] < 1].copy()
    if sure_bets.empty:
        logging.info("No sure bets found.")
        return df_merged, pd.DataFrame()

    # Calculate stake distribution
    sure_bets[['Stake1','Stake2','Stake3','Profit1','Profit2','Profit3']] = sure_bets.apply(
        lambda row: pd.Series(beat_bookies(row['best_1'], row['best_x'], row['best_2'], total_stake)),
        axis=1
    )

    logging.info(f"Sure bets found: {len(sure_bets)}")
    return df_merged, sure_bets

# ---------------------------------------------------- #
#              Archiving & Housekeeping               #
# ---------------------------------------------------- #
def archive_file(file_path: str, archive_folder: str = ARCHIVE_FOLDER):
    """Zip up and move the file to archive with a timestamped name."""
    try:
        os.makedirs(archive_folder, exist_ok=True)
        base_name = os.path.basename(file_path)
        name, ext = os.path.splitext(base_name)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        new_name = f"{name}_{timestamp}{ext}"
        new_path = os.path.join(os.path.dirname(file_path), new_name)

        os.rename(file_path, new_path)
        zip_name = f"{name}_{timestamp}.zip"
        zip_path = os.path.join(os.path.dirname(file_path), zip_name)
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(new_path, arcname=new_name)
        shutil.move(zip_path, archive_folder)
        os.remove(new_path)
        logging.info(f"Archived file {file_path} -> {os.path.join(archive_folder, zip_name)}")
    except Exception as e:
        logging.error(f"Failed to archive {file_path}: {e}")

# ---------------------------------------------------- #
#                       Main                          #
# ---------------------------------------------------- #
def main():
    # 1) Load data (in parallel)
    pickle_bookies = load_pickle_data_concurrent(PICKLE_FOLDER_PATH)
    excel_bookies = load_excel_data_concurrent(EXCEL_FOLDER_PATH)
    all_bookies = {**pickle_bookies, **excel_bookies}

    if not all_bookies:
        logging.error("No data loaded from pickle or Excel.")
        print("No data loaded. Check logs for details.")
        return

    # 2) Clean each DataFrame
    cleaned_bookies = {}
    for bookie_name, df in all_bookies.items():
        try:
            df_clean = extract_relevant_columns(df, bookie_name)
            if not df_clean.empty:
                cleaned_bookies[bookie_name] = df_clean
        except Exception as e:
            logging.error(f"Failed to clean {bookie_name}: {e}")

    # If fewer than 2 remain, no merges are possible
    if len(cleaned_bookies) < 2:
        logging.error("Fewer than 2 valid bookies after cleaning.")
        print("Fewer than 2 valid bookies after cleaning. Check logs.")
        return

    # 3) Multi-way merge
    df_merged_all = multiway_fuzzy_merge(cleaned_bookies, threshold=FUZZY_THRESHOLD)
    if df_merged_all.empty:
        logging.error("No merges produced a valid DataFrame.")
        print("No merges produced a valid DataFrame. Check logs.")
        return

    # 4) Calculate implied probabilities & sure bets
    master_df, sure_bets = calculate_implied_probabilities(df_merged_all, TOTAL_STAKE)

    # 5) If sure bets exist, save them
    if not sure_bets.empty:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Sure Bets Summary
        sheet_summary = wb.create_sheet("Sure Bets Summary")
        headers_summary = [
            "Home","Time",
            "Best_1","Best_1_Source",
            "Best_X","Best_X_Source",
            "Best_2","Best_2_Source",
            "Implied Probability"
        ]
        sheet_summary.append(headers_summary)
        for _, row in sure_bets.iterrows():
            sheet_summary.append([
                str(row['home']),
                str(row['time']),
                row['best_1'],
                row['best_1_source'],
                row['best_x'],
                row['best_x_source'],
                row['best_2'],
                row['best_2_source'],
                row['implied']
            ])

        # Detailed Staking
        sheet_detail = wb.create_sheet("Detailed Staking")
        headers_detail = [
            "Home","Time",
            "Best_1","Best_1_Source",
            "Best_X","Best_X_Source",
            "Best_2","Best_2_Source",
            "Stake1","Stake2","Stake3",
            "Profit1","Profit2","Profit3"
        ]
        sheet_detail.append(headers_detail)
        for _, row in sure_bets.iterrows():
            sheet_detail.append([
                str(row['home']),
                str(row['time']),
                row['best_1'],
                row['best_1_source'],
                row['best_x'],
                row['best_x_source'],
                row['best_2'],
                row['best_2_source'],
                row['Stake1'],
                row['Stake2'],
                row['Stake3'],
                row['Profit1'],
                row['Profit2'],
                row['Profit3']
            ])

        result_path = os.path.join(OUTPUT_FOLDER, "sure_bets_strict.xlsx")
        try:
            wb.save(result_path)
            logging.info(f"Sure bets saved to {result_path}")
            print(f"Sure bets saved to {result_path}")
        except Exception as e:
            logging.error(f"Failed to save sure bets: {e}")
            print("Failed to save sure bets. Check logs.")
    else:
        logging.info("No sure bets found after calculations.")

    # 6) Archive processed files
    # Archive .pkl
    for file_path in glob.glob(os.path.join(PICKLE_FOLDER_PATH, "*.pkl")):
        archive_file(file_path, ARCHIVE_FOLDER)

    # Archive .xlsx
    try:
        for file_name in os.listdir(EXCEL_FOLDER_PATH):
            if file_name.lower().endswith('.xlsx'):
                file_path = os.path.join(EXCEL_FOLDER_PATH, file_name)
                archive_file(file_path, ARCHIVE_FOLDER)
    except FileNotFoundError:
        logging.error(f"Excel folder not found: {EXCEL_FOLDER_PATH}")

if __name__ == "__main__":
    main()
