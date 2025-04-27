import glob
import os
import pickle
import pandas as pd
from fuzzywuzzy import process, fuzz
import openpyxl
import itertools
import logging

# Configure logging
logging.basicConfig(
    filename='log/find_surebet.log',
    filemode='w',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# ---------- CONFIG ----------------- #
FOLDER_PATH = "./pickle_data"  # Folder containing .pickle files
FUZZY_THRESHOLD = 80
TOTAL_STAKE = 100
# ----------------------------------- #

def load_bookie_data(folder_path: str) -> dict:
    """
    Loads all .pickle files from the folder_path into a dictionary { bookie_name: df }.
    """
    bookies = {}
    for file_path in glob.glob(os.path.join(folder_path, "*")):
        file_name = os.path.basename(file_path)
        bookie_name = os.path.splitext(file_name)[0]
        try:
            df = pickle.load(open(file_path, "rb"))
            bookies[bookie_name] = df
        except Exception as e:
            logging.error(f"Could not load {file_path}. Error: {e}")
    return bookies

def extract_relevant_columns(df, bookmaker_name):
    """
    Ensures we only keep the columns: ['home', '1', 'x', '2'] 
    and that they're numeric where appropriate.
    """
    needed_cols = ['home', '1', 'x', '2']
    missing_cols = [col for col in needed_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(
            f"DataFrame for {bookmaker_name} is missing columns: {missing_cols}"
        )
    df = df[needed_cols].copy()
    # Convert odds to numeric
    for col in ['1', 'x', '2']:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    # Drop any rows that have NaN in these essential columns
    df.dropna(subset=['home', '1', 'x', '2'], inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df

def fuzzy_merge_two_dfs(df_master, df_new, master_key='home', new_key='home',
                       threshold=80, scorer=fuzz.token_set_ratio, new_bookie_name=None):
    """
    Fuzzy-merge df_new into df_master, matching by 'home'.
    - df_master has columns: home, plus possibly other bookie_1, bookie_x, bookie_2.
    - df_new has columns: home, 1, x, 2.
    
    We add new columns to df_master: [new_bookie_name+'_1', new_bookie_name+'_x', new_bookie_name+'_2'].
    """

    # Convert the 'home' column of df_new to list for fuzzy matching
    df_new_list = df_new[new_key].tolist()

    # We will build a list of rows for the new merged master
    merged_rows = []

    # For each row in the current master, find best fuzzy match in df_new
    for _, row_m in df_master.iterrows():
        home_master = row_m[master_key]
        match = process.extractOne(home_master, df_new_list, scorer=scorer)
        if match:
            best_match_str, best_score = match
            if best_score >= threshold:
                # get the row(s) in df_new that match best_match_str
                possible_matches = df_new[df_new[new_key] == best_match_str]
                # In theory, there should be only one or just a few matches
                for _, row_n in possible_matches.iterrows():
                    # Copy the original master row
                    combined = row_m.to_dict()
                    # Add new columns with the new bookie’s odds
                    combined[f"{new_bookie_name}_1"] = row_n['1']
                    combined[f"{new_bookie_name}_x"] = row_n['x']
                    combined[f"{new_bookie_name}_2"] = row_n['2']
                    merged_rows.append(combined)
            else:
                # No good match found, just carry the master row with N/A for new columns
                combined = row_m.to_dict()
                combined[f"{new_bookie_name}_1"] = None
                combined[f"{new_bookie_name}_x"] = None
                combined[f"{new_bookie_name}_2"] = None
                merged_rows.append(combined)
        else:
            # No match at all
            combined = row_m.to_dict()
            combined[f"{new_bookie_name}_1"] = None
            combined[f"{new_bookie_name}_x"] = None
            combined[f"{new_bookie_name}_2"] = None
            merged_rows.append(combined)

    # Convert list of dicts back to a DataFrame
    df_merged = pd.DataFrame(merged_rows)
    return df_merged

def multiway_fuzzy_merge(cleaned_bookies: dict, threshold=80):
    """
    Iteratively fuzzy merges all bookies into one wide DataFrame.
    """
    # Turn dict into list of (bookie_name, df)
    bookie_items = list(cleaned_bookies.items())

    # Start with the first as 'df_master'
    base_bookie_name, df_master = bookie_items[0]
    # Rename its '1','x','2' columns to {base_bookie_name}_1, etc.
    df_master = df_master.rename(columns={
        '1': f"{base_bookie_name}_1",
        'x': f"{base_bookie_name}_x",
        '2': f"{base_bookie_name}_2",
    })

    # For each subsequent bookie, fuzzy-merge into df_master
    for bookie_name, df_new in bookie_items[1:]:
        df_master = fuzzy_merge_two_dfs(
            df_master, 
            df_new, 
            threshold=threshold, 
            scorer=fuzz.token_set_ratio, 
            new_bookie_name=bookie_name
        )

    return df_master

def beat_bookies(odds1, odds2, odds3, total_stake=100):
    """
    Splits total_stake among 3 outcomes so that any winning outcome yields the same profit.
    If no solution (or negative stakes), returns 0's.
    """
    from sympy import symbols, Eq, solve
    x, y, z = symbols('x y z', real=True, nonnegative=True)
    eq1 = Eq(x + y + z, total_stake)
    # We want odds2*y - odds1*x = 0 => odds2*y = odds1*x => ratio
    eq2 = Eq((odds2 * y) - (odds1 * x), 0)
    eq3 = Eq((odds3 * z) - (odds1 * x), 0)
    sol = solve((eq1, eq2, eq3), (x, y, z), dict=True)
    if not sol:
        return {
            'Stake1': 0, 'Stake2': 0, 'Stake3': 0,
            'Profit1': 0, 'Profit2': 0, 'Profit3': 0
        }
    s = sol[0]
    stake1, stake2, stake3 = s[x], s[y], s[z]
    return {
        'Stake1': stake1,
        'Stake2': stake2,
        'Stake3': stake3,
        'Profit1': odds1 * stake1 - total_stake,
        'Profit2': odds2 * stake2 - total_stake,
        'Profit3': odds3 * stake3 - total_stake
    }

def safe_format(value):
    if isinstance(value, (float, int)):
        return f"{value:.4f}"
    return str(value)

def main():
    # 1) Load all data
    all_bookies = load_bookie_data(FOLDER_PATH)
    if not all_bookies:
        logging.error("No .pickle files found or none could be loaded. Exiting.")
        return
    
    # 2) Clean each DataFrame
    cleaned_bookies = {}
    for bookie_name, df in all_bookies.items():
        try:
            df_clean = extract_relevant_columns(df, bookie_name)
            cleaned_bookies[bookie_name] = df_clean
        except ValueError as e:
            logging.error(f"Skipping {bookie_name}. Error: {e}")

    # If fewer than 2 valid bookies remain, no merges are possible
    if len(cleaned_bookies) < 2:
        logging.error("Fewer than 2 valid bookies after cleaning. Exiting.")
        return

    # 3) Multi-way fuzzy merge: produce one wide DataFrame
    df_merged_all = multiway_fuzzy_merge(cleaned_bookies, threshold=FUZZY_THRESHOLD)
    if df_merged_all.empty:
        logging.error("No merges could be done. Exiting.")
        return

    # Now df_merged_all should look like:
    # home, BookieA_1, BookieA_x, BookieA_2, BookieB_1, BookieB_x, BookieB_2, ...

    # 4) Identify columns that correspond to odds for each bookie
    all_cols = df_merged_all.columns
    # We skip 'home' (and fuzzy_score if it exists), keep only columns that end with _1, _x, _2
    odds_1_cols = [c for c in all_cols if c.endswith("_1")]
    odds_x_cols = [c for c in all_cols if c.endswith("_x")]
    odds_2_cols = [c for c in all_cols if c.endswith("_2")]

    # 5) For each row, find the best odds across all bookies
    df_merged_all['best_1'] = df_merged_all[odds_1_cols].max(axis=1, skipna=True)
    df_merged_all['best_x'] = df_merged_all[odds_x_cols].max(axis=1, skipna=True)
    df_merged_all['best_2'] = df_merged_all[odds_2_cols].max(axis=1, skipna=True)

    # 5a) Also find which bookie gave the best odds
    # idxmax() returns the column name that has the highest value along the row
    df_merged_all['best_1_source'] = df_merged_all[odds_1_cols].idxmax(axis=1)
    df_merged_all['best_x_source'] = df_merged_all[odds_x_cols].idxmax(axis=1)
    df_merged_all['best_2_source'] = df_merged_all[odds_2_cols].idxmax(axis=1)

    # For readability, remove the trailing "_1" or "_x" or "_2" from the column name
    df_merged_all['best_1_source'] = df_merged_all['best_1_source'].apply(
        lambda x: x[:-2] if isinstance(x, str) and x.endswith('_1') else x)
    df_merged_all['best_x_source'] = df_merged_all['best_x_source'].apply(
        lambda x: x[:-2] if isinstance(x, str) and x.endswith('_x') else x)
    df_merged_all['best_2_source'] = df_merged_all['best_2_source'].apply(
        lambda x: x[:-2] if isinstance(x, str) and x.endswith('_2') else x)

    # 6) Calculate the implied probability
    df_merged_all['implied'] = (1 / df_merged_all['best_1']) + \
                               (1 / df_merged_all['best_x']) + \
                               (1 / df_merged_all['best_2'])

    # 7) Filter to see which are sure bets (implied < 1)
    sure_bets = df_merged_all[df_merged_all['implied'] < 1].copy()
    sure_bets.reset_index(drop=True, inplace=True)

    if sure_bets.empty:
        logging.error("No sure bets found.")
        return

    # 8) Write to Excel: summary + detailed
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb["Sheet"] if "Sheet" in wb.sheetnames else None
    if default_sheet:
        wb.remove(default_sheet)

    # Sure Bets Summary
    summary_sheet = wb.create_sheet(title="Sure Bets Summary")
    summary_sheet_headers = [
        "Home", 
        "best_1", "best_1_source",
        "best_x", "best_x_source",
        "best_2", "best_2_source",
        "implied"
    ]
    summary_sheet.append(summary_sheet_headers)

    for _, row in sure_bets.iterrows():
        summary_sheet.append([
            safe_format(row['home']),
            safe_format(row['best_1']),
            safe_format(row['best_1_source']),
            safe_format(row['best_x']),
            safe_format(row['best_x_source']),
            safe_format(row['best_2']),
            safe_format(row['best_2_source']),
            safe_format(row['implied']),
        ])

    # Detailed sheet with staking plan
    detail_sheet = wb.create_sheet(title="Detailed Staking")
    detail_sheet_headers = [
        "Home", 
        "best_1", "best_1_source",
        "best_x", "best_x_source",
        "best_2", "best_2_source",
        "Stake1", "Stake2", "Stake3",
        "Profit1", "Profit2", "Profit3",
    ]
    detail_sheet.append(detail_sheet_headers)

    for _, row in sure_bets.iterrows():
        odds1 = row['best_1']
        odds2 = row['best_x']
        odds3 = row['best_2']
        plan = beat_bookies(odds1, odds2, odds3, TOTAL_STAKE)
        detail_sheet.append([
            safe_format(row['home']),
            safe_format(odds1),
            safe_format(row['best_1_source']),
            safe_format(odds2),
            safe_format(row['best_x_source']),
            safe_format(odds3),
            safe_format(row['best_2_source']),
            safe_format(plan['Stake1']),
            safe_format(plan['Stake2']),
            safe_format(plan['Stake3']),
            safe_format(plan['Profit1']),
            safe_format(plan['Profit2']),
            safe_format(plan['Profit3']),
        ])

    wb.save("results/sure_bets.xlsx")
    logging.info("Sure bet results saved to results/sure_bets.xlsx")


if __name__ == "__main__":
    main()
