import argparse
import glob
import os
import shutil
import pickle
import pandas as pd
import openpyxl
import logging
from sympy import symbols, Eq, solve
import datetime
import zipfile
from typing import Dict, List, Tuple, Optional
import yaml  # pip install pyyaml
import numpy as np

# Use RapidFuzz instead of fuzzywuzzy
from rapidfuzz import process, fuzz

# ------------------------------------- #
#          Custom Exceptions           #
# ------------------------------------- #
class MissingColumnError(Exception):
    """Raised when a required column is missing in a DataFrame."""
    pass

class InvalidOddsError(Exception):
    """Raised when a DataFrame has invalid or extreme odds."""
    pass

# ------------------------------------- #
#       Load/Parse External Config      #
# ------------------------------------- #
def load_config(config_path: str = "config/config.yaml") -> dict:
    default_config = {
        "PICKLE_FOLDER_PATH": "./pickle_data",
        "EXCEL_FOLDER_PATH": "./data",
        "FUZZY_THRESHOLD": 80,
        "TOTAL_STAKE": 100,
        "OUTPUT_FOLDER": "./results",
        "ARCHIVE_FOLDER": "./archive",
        "MIN_ODDS": 1.01,
        "MAX_ODDS": 100.0
    }
    if not os.path.exists(config_path):
        logging.warning(f"Config file '{config_path}' not found. Using default config.")
        return default_config
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            user_config = yaml.safe_load(f) or {}
        for key, val in default_config.items():
            if key not in user_config:
                user_config[key] = val
        return user_config
    except Exception as e:
        logging.error(f"Failed to read config file '{config_path}': {e}")
        return default_config

def parse_arguments(config: dict) -> dict:
    parser = argparse.ArgumentParser(description="Sure Bet Script with Enhanced Features")
    parser.add_argument(
        "--threshold",
        type=int,
        default=config["FUZZY_THRESHOLD"],
        help="Fuzzy matching threshold (default from config)."
    )
    parser.add_argument(
        "--total-stake",
        type=float,
        default=config["TOTAL_STAKE"],
        help="Total stake for each sure bet calculation (default from config)."
    )
    args = parser.parse_args()
    config["FUZZY_THRESHOLD"] = args.threshold
    config["TOTAL_STAKE"]     = args.total_stake
    return config

# ------------------------------------- #
#        Configuration & Logging        #
# ------------------------------------- #
os.makedirs("log", exist_ok=True)
logging.basicConfig(
    filename='log/sure_bet_strict.log',
    filemode='w',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Now require both teams!
REQUIRED_COLUMNS = ['home', 'away', '1', 'x', '2', 'time']

TEAM_NAME_ALIASES = {
    # "Man Utd": "Manchester United",
}

EXCLUDED_KEYWORDS = ["betlive", "betole", "sportplus"]

# ------------------------------------- #
#         Helper / Utility Code         #
# ------------------------------------- #
def alias_team_name(name: str) -> str:
    return TEAM_NAME_ALIASES.get(name, name)

def validate_schema(df: pd.DataFrame, bookie_name: str) -> None:
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        logging.error(f"{bookie_name}: Missing columns {missing}")
        raise MissingColumnError(f"{bookie_name}: Missing required columns {missing}")

def filter_outliers(df: pd.DataFrame, bookie_name: str, min_odds: float, max_odds: float) -> None:
    before = len(df)
    cond = (
        (df['1'] < min_odds) | (df['1'] > max_odds) |
        (df['x'] < min_odds) | (df['x'] > max_odds) |
        (df['2'] < min_odds) | (df['2'] > max_odds)
    )
    df.drop(df[cond].index, inplace=True)
    after = len(df)
    if after < before:
        logging.warning(f"{bookie_name}: Dropped {before - after} rows with outlier odds.")

def convert_time_to_utc(df: pd.DataFrame, bookie_name: str, time_col: str = 'time') -> None:
    if time_col not in df.columns:
        return
    try:
        # Optionally, if you know the format (e.g. '%Y-%m-%d %H:%M:%S'), specify it here.
        df.loc[:, time_col] = pd.to_datetime(df.loc[:, time_col], errors='coerce').dt.tz_localize(None)
        logging.debug(f"{bookie_name}: Time column standardized to naive datetime.")
    except Exception as e:
        logging.error(f"{bookie_name}: Error converting time to datetime. {e}")

# ------------------------------------- #
#        File Loading and Cleaning      #
# ------------------------------------- #
def load_pickle_data(folder_path: str) -> dict:
    bookies = {}
    for file_path in glob.glob(os.path.join(folder_path, "*.pkl")):
        file_name = os.path.basename(file_path)
        bookie_name = os.path.splitext(file_name)[0]
        try:
            with open(file_path, "rb") as f:
                df = pickle.load(f)
            logging.info(f"Loaded pickle file: {file_name}")
            bookies[bookie_name] = df
        except Exception as e:
            logging.error(f"Could not load {file_path}. Error: {e}")
    return bookies

def load_excel_data(folder_path: str) -> dict:
    bookies = {}
    for file_path in glob.glob(os.path.join(folder_path, "*.xlsx")) + glob.glob(os.path.join(folder_path, "*.XLSX")):
        file_name = os.path.basename(file_path)
        bookie_name = os.path.splitext(file_name)[0]
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            logging.info(f"Loaded Excel file: {file_name}")
            bookies[bookie_name] = df
        except Exception as e:
            logging.error(f"Could not load {file_path}. Error: {e}")
    return bookies

def standardize_columns(df: pd.DataFrame, bookie_name: str) -> pd.DataFrame:
    column_mappings = {
        'meridianbet_fudbal': {
            'home': 'home',
            'away': 'away',
            '1': '1',
            '2': '2',
            'Oba tima postižu bar po 1 gol i ukupno golova na meču manje od 3': 'x_extra',
            'Venados FC': 'irrelevant_column'
        },
        'sportplus': {
            'kickoff': 'time',
            'home': 'home',
            'away': 'away',
            '1.00': '1',
            'x': 'x',
            '2.00': '2',
            'time_duplicate': 'duplicate_time'
        },
    }
    applied_mapping = {}
    for key in column_mappings:
        if key.lower() in bookie_name.lower():
            applied_mapping = column_mappings[key]
            break

    if not applied_mapping:
        possible_cols = df.columns.str.lower()
        if 'home' in possible_cols:
            df.rename(columns={'home':'home'}, inplace=True, errors='ignore')
        if 'away' in possible_cols:
            df.rename(columns={'away':'away'}, inplace=True, errors='ignore')
        if '1' in possible_cols:
            df.rename(columns={'1':'1'}, inplace=True, errors='ignore')
        if 'x' in possible_cols:
            df.rename(columns={'x':'x'}, inplace=True, errors='ignore')
        if '2' in possible_cols:
            df.rename(columns={'2':'2'}, inplace=True, errors='ignore')
        if 'time' in possible_cols:
            df.rename(columns={'time':'time'}, inplace=True, errors='ignore')
        if 'kickoff' in possible_cols:
            df.rename(columns={'kickoff':'time'}, inplace=True, errors='ignore')
    else:
        df.rename(columns=applied_mapping, inplace=True)
    if 'irrelevant_column' in applied_mapping:
        df.drop(columns=['irrelevant_column'], inplace=True, errors='ignore')
    if 'x_extra' in applied_mapping:
        df.drop(columns=['x_extra'], inplace=True, errors='ignore')
    if 'duplicate_time' in applied_mapping:
        df.drop(columns=['duplicate_time'], inplace=True, errors='ignore')
    if df.columns.duplicated().any():
        logging.error(f"{bookie_name}: Duplicate columns found after standardization.")
        df = df.loc[:, ~df.columns.duplicated()]
    return df

def extract_relevant_columns(df: pd.DataFrame,
                             bookie_name: str,
                             min_odds: float,
                             max_odds: float) -> pd.DataFrame:
    df = standardize_columns(df, bookie_name)
    validate_schema(df, bookie_name)
    for col in ['1', 'x', '2']:
        df.loc[:, col] = pd.to_numeric(df.loc[:, col].astype(str).str.replace(',', '.'), errors='coerce')
    convert_time_to_utc(df, bookie_name, time_col='time')
    if 'away' in df.columns:
        df.loc[:, 'away'] = df.loc[:, 'away'].astype(str).apply(alias_team_name)
    df['sum_odds'] = df['1'] + df['x'] + df['2']
    df.sort_values(by='sum_odds', inplace=True)
    df = df.groupby(["home", "away", "time"], as_index=False).first()
    df.drop(columns=['sum_odds'], inplace=True, errors='ignore')
    before_drop = len(df)
    df.dropna(subset=['home','away','1','x','2','time'], inplace=True)
    df = df[(df['1'] > 0) & (df['x'] > 0) & (df['2'] > 0)]
    after_drop = len(df)
    if after_drop < before_drop:
        logging.info(f"{bookie_name}: Dropped {before_drop - after_drop} rows with NaN or invalid odds.")
    filter_outliers(df, bookie_name, min_odds, max_odds)
    df.loc[:, 'home'] = df.loc[:, 'home'].astype(str).apply(alias_team_name)
    df.reset_index(drop=True, inplace=True)
    if df.empty:
        logging.warning(f"{bookie_name}: DataFrame is empty after cleaning.")
    return df

# --------------------------------------------------- #
#   Full Multiway Fuzzy Merge (Using RapidFuzz & cdist)  #
# --------------------------------------------------- #
def full_multiway_fuzzy_merge(cleaned_bookies: Dict[str, pd.DataFrame], threshold=80) -> pd.DataFrame:
    """
    Build a master set of merge keys from all bookies.
    For each file, use rapidfuzz.process.cdist to quickly get match scores
    between the master key list and that file's keys.
    """
    master_keys = set()
    keys_by_bookie = {}
    rows_by_bookie = {}
    # Add a merge key column to every file and collect keys
    for bookie, df in cleaned_bookies.items():
        df['merge_key'] = df.apply(lambda r: f"{r['home'].strip()} vs {r['away'].strip()} {r['time']}", axis=1)
        keys = df['merge_key'].tolist()
        keys_by_bookie[bookie] = keys
        master_keys.update(keys)
        # Index the DataFrame by merge_key for quick lookup
        rows_by_bookie[bookie] = df.set_index('merge_key', drop=False)
    master_keys = list(master_keys)
    master_keys_array = np.array(master_keys)

    # For each bookie, vectorized matching: compute cdist between master_keys and file's keys.
    results_by_bookie = {}
    for bookie, keys_list in keys_by_bookie.items():
        # Compute distances for all master keys against this file's keys.
        # processor=None here because our keys are already strings.
        scores = process.cdist(master_keys, keys_list, scorer=fuzz.token_set_ratio, processor=lambda x: x)
        scores = np.array(scores)  # shape: (len(master_keys), len(keys_list))
        best_idx = np.argmax(scores, axis=1)
        best_score = np.max(scores, axis=1)
        # Build a dictionary mapping master key -> matched key if score>=threshold, else None
        bookie_matches = {}
        for i, mk in enumerate(master_keys):
            if best_score[i] >= threshold:
                bookie_matches[mk] = keys_list[best_idx[i]]
            else:
                bookie_matches[mk] = None
        results_by_bookie[bookie] = bookie_matches

    master_rows = []
    # For each master key, compile the row from each file.
    for mk in master_keys:
        row = {}
        # Also try to fill base information from the first file that has a match.
        base_filled = False
        for bookie in cleaned_bookies.keys():
            match_key = results_by_bookie[bookie].get(mk)
            if match_key is not None:
                # Get row from this file
                matched_row = rows_by_bookie[bookie].loc[match_key].to_dict()
                row[f"{bookie}_1"] = matched_row.get('1')
                row[f"{bookie}_x"] = matched_row.get('x')
                row[f"{bookie}_2"] = matched_row.get('2')
                if not base_filled:
                    row['home'] = matched_row.get('home')
                    row['away'] = matched_row.get('away')
                    row['time'] = matched_row.get('time')
                    base_filled = True
            else:
                row[f"{bookie}_1"] = None
                row[f"{bookie}_x"] = None
                row[f"{bookie}_2"] = None
        master_rows.append(row)
    master_df = pd.DataFrame(master_rows)
    return master_df

def beat_bookies(odds1, odds2, odds3, total_stake=100):
    try:
        x, y, z = symbols('x y z', real=True, nonnegative=True)
        eq1 = Eq(x + y + z, total_stake)
        eq2 = Eq(odds1 * x, odds2 * y)
        eq3 = Eq(odds1 * x, odds3 * z)
        sol = solve((eq1, eq2, eq3), (x, y, z), dict=True)
        if not sol:
            return {'Stake1': 0, 'Stake2': 0, 'Stake3': 0,
                    'Profit1': 0, 'Profit2': 0, 'Profit3': 0}
        s = sol[0]
        stake1, stake2, stake3 = s[x], s[y], s[z]
        if stake1 < 0 or stake2 < 0 or stake3 < 0:
            return {'Stake1': 0, 'Stake2': 0, 'Stake3': 0,
                    'Profit1': 0, 'Profit2': 0, 'Profit3': 0}
        return {
            'Stake1': float(stake1),
            'Stake2': float(stake2),
            'Stake3': float(stake3),
            'Profit1': float(odds1 * stake1 - total_stake),
            'Profit2': float(odds2 * stake2 - total_stake),
            'Profit3': float(odds3 * stake3 - total_stake)
        }
    except Exception as e:
        logging.error(f"Error in beat_bookies calculation: {e}")
        return {'Stake1': 0, 'Stake2': 0, 'Stake3': 0,
                'Profit1': 0, 'Profit2': 0, 'Profit3': 0}

def safe_format(value):
    if isinstance(value, (float, int)):
        return f"{value:.4f}"
    return str(value)

# ------------------------------------- #
#        Archiving & Housekeeping       #
# ------------------------------------- #
def archive_file(file_path: str, archive_folder: str):
    try:
        date_str = datetime.datetime.now().strftime("%Y%m%d")
        daily_archive_folder = os.path.join(archive_folder, date_str)
        os.makedirs(daily_archive_folder, exist_ok=True)
        base_name = os.path.basename(file_path)
        name, ext = os.path.splitext(base_name)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        new_name = f"{name}_{timestamp}{ext}"
        new_path = os.path.join(os.path.dirname(file_path), new_name)
        os.rename(file_path, new_path)
        logging.info(f"Renamed file {file_path} to {new_path}")
        zip_name = f"{name}_{timestamp}.zip"
        zip_path = os.path.join(os.path.dirname(file_path), zip_name)
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(new_path, arcname=new_name)
        logging.info(f"Created zip archive: {zip_path}")
        shutil.move(zip_path, daily_archive_folder)
        logging.info(f"Moved zip archive to {daily_archive_folder}: {zip_path}")
        os.remove(new_path)
        logging.info(f"Deleted original renamed file: {new_path}")
    except Exception as e:
        logging.error(f"Failed to archive {file_path}. Error: {e}")

def identify_matched_unmatched(master_df: pd.DataFrame, target_bookie: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    col_prefix_1 = f"{target_bookie}_1"
    matched = master_df[~master_df[col_prefix_1].isna()].copy()
    unmatched = master_df[master_df[col_prefix_1].isna()].copy()
    logging.info(f"{target_bookie}: Total matches attempted: {len(master_df)}")
    logging.info(f"{target_bookie}: Successfully merged matches: {len(matched)}")
    logging.info(f"{target_bookie}: Unmerged matches: {len(unmatched)}")
    if not unmatched.empty:
        logging.info(f"Detailed unmatched {target_bookie} matches:")
        for idx, row in unmatched.iterrows():
            home_val = row.get('home','N/A')
            away_val = row.get('away','N/A')
            time_val = row.get('time','N/A')
            logging.info(f"Unmatched {idx}: Home='{home_val}', Away='{away_val}', Time='{time_val}'")
    else:
        logging.info(f"All matches from {target_bookie} were successfully merged.")
    return matched, unmatched

def calculate_implied_probabilities(master_df: pd.DataFrame, total_stake: float) -> Tuple[pd.DataFrame, pd.DataFrame]:
    all_cols = master_df.columns
    odds_1_cols = [c for c in all_cols if c.endswith("_1")]
    odds_x_cols = [c for c in all_cols if c.endswith("_x")]
    odds_2_cols = [c for c in all_cols if c.endswith("_2")]
    if not odds_1_cols or not odds_x_cols or not odds_2_cols:
        logging.error("Missing odds columns after merging. Exiting.")
        print("Missing odds columns after merging. Check 'log/sure_bet_strict.log' for details.")
        return master_df, pd.DataFrame()
    master_df['best_1'] = master_df[odds_1_cols].min(axis=1, skipna=True)
    master_df['best_x'] = master_df[odds_x_cols].min(axis=1, skipna=True)
    master_df['best_2'] = master_df[odds_2_cols].min(axis=1, skipna=True)
    try:
        master_df['best_1_source'] = master_df[odds_1_cols].idxmin(axis=1).str.replace('_1','')
        master_df['best_x_source'] = master_df[odds_x_cols].idxmin(axis=1).str.replace('_x','')
        master_df['best_2_source'] = master_df[odds_2_cols].idxmin(axis=1).str.replace('_2','')
    except AttributeError as e:
        logging.error(f"Error identifying best odds sources: {e}")
        master_df['best_1_source'] = None
        master_df['best_x_source'] = None
        master_df['best_2_source'] = None
    df_valid = master_df[
        (master_df['best_1'] > 0) &
        (master_df['best_x'] > 0) &
        (master_df['best_2'] > 0)
    ].copy()
    excluded = len(master_df) - len(df_valid)
    if excluded > 0:
        logging.warning(f"Excluded {excluded} rows with zero or missing odds.")
    try:
        df_valid['implied'] = (1 / df_valid['best_1']) + \
                              (1 / df_valid['best_x']) + \
                              (1 / df_valid['best_2'])
    except ZeroDivisionError as e:
        logging.error(f"ZeroDivisionError in implied probability calculation: {e}")
        df_valid = df_valid[(df_valid['best_1']>0)&(df_valid['best_x']>0)&(df_valid['best_2']>0)].copy()
        df_valid['implied'] = (1 / df_valid['best_1']) + \
                              (1 / df_valid['best_x']) + \
                              (1 / df_valid['best_2'])
    sure_bets = df_valid[df_valid['implied'] < 1].copy()
    sure_bets.reset_index(drop=True, inplace=True)
    if sure_bets.empty:
        logging.error("No sure bets found.")
        print("No sure bets found.")
        return master_df, pd.DataFrame()
    logging.info(f"Number of sure bets found: {len(sure_bets)}")
    sure_bets[['Stake1','Stake2','Stake3','Profit1','Profit2','Profit3']] = sure_bets.apply(
        lambda row: pd.Series(beat_bookies(row['best_1'], row['best_x'], row['best_2'], total_stake)),
        axis=1
    )
    return master_df, sure_bets

# ------------------------------------- #
#                Main Script           #
# ------------------------------------- #
def main():
    config = load_config("config/config.yaml")
    config = parse_arguments(config)
    PICKLE_FOLDER_PATH = config["PICKLE_FOLDER_PATH"]
    EXCEL_FOLDER_PATH  = config["EXCEL_FOLDER_PATH"]
    FUZZY_THRESHOLD    = config["FUZZY_THRESHOLD"]
    TOTAL_STAKE        = config["TOTAL_STAKE"]
    OUTPUT_FOLDER      = config["OUTPUT_FOLDER"]
    ARCHIVE_FOLDER     = config["ARCHIVE_FOLDER"]
    MIN_ODDS           = config["MIN_ODDS"]
    MAX_ODDS           = config["MAX_ODDS"]

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    os.makedirs(ARCHIVE_FOLDER, exist_ok=True)

    pickle_bookies = load_pickle_data(PICKLE_FOLDER_PATH)
    if not pickle_bookies:
        logging.warning(f"No .pkl files found in {PICKLE_FOLDER_PATH}. Falling back to Excel in ./data.")
        EXCEL_FOLDER_PATH = "./data"
        excel_bookies = load_excel_data(EXCEL_FOLDER_PATH)
        all_bookies   = excel_bookies
    else:
        excel_bookies = load_excel_data(EXCEL_FOLDER_PATH)
        all_bookies   = {**pickle_bookies, **excel_bookies}

    if not all_bookies:
        logging.error("No data files loaded (neither pkl nor excel). Exiting.")
        print("No data files loaded. Check 'log/sure_bet_strict.log' for details.")
        return

    cleaned_bookies = {}
    for bookie_name, df in all_bookies.items():
        try:
            df_clean = extract_relevant_columns(
                df, 
                bookie_name,
                min_odds=MIN_ODDS,
                max_odds=MAX_ODDS
            )
            if df_clean.empty:
                logging.warning(f"{bookie_name}: DataFrame is empty after cleaning; skipping.")
                continue
            cleaned_bookies[bookie_name] = df_clean
            logging.info(f"Cleaned data for bookie: {bookie_name}")
        except MissingColumnError as e:
            logging.error(f"Skipping {bookie_name} due to missing columns: {e}")
        except InvalidOddsError as e:
            logging.error(f"Skipping {bookie_name} due to invalid odds: {e}")
        except Exception as e:
            logging.error(f"Unexpected error cleaning {bookie_name}: {e}")

    if len(cleaned_bookies) < 2:
        logging.error("Fewer than 2 valid bookies after cleaning. Exiting.")
        print("Fewer than 2 valid bookies after cleaning. Check 'log/sure_bet_strict.log' for details.")
        return

    # Use full multiway merge that attempts to match every file with every game.
    master_df = full_multiway_fuzzy_merge(cleaned_bookies, threshold=FUZZY_THRESHOLD)
    if master_df.empty:
        logging.error("No merged data produced. Exiting.")
        print("No merges could be done. Check 'log/sure_bet_strict.log' for details.")
        return

    master_df, sure_bets = calculate_implied_probabilities(master_df, TOTAL_STAKE)
    if not sure_bets.empty:
        wb = openpyxl.Workbook()
        default_sheet = wb["Sheet"] if "Sheet" in wb.sheetnames else None
        if default_sheet:
            wb.remove(default_sheet)
        summary_sheet = wb.create_sheet(title="Sure Bets Summary")
        summary_headers = [
            "Home", "Away", "Time",
            "Best_1", "Best_1_Source",
            "Best_X", "Best_X_Source",
            "Best_2", "Best_2_Source",
            "Implied Probability"
        ]
        summary_sheet.append(summary_headers)
        for _, row in sure_bets.iterrows():
            summary_sheet.append([
                safe_format(row['home']),
                safe_format(row['away']),
                safe_format(row['time']),
                safe_format(row['best_1']),
                safe_format(row['best_1_source']),
                safe_format(row['best_x']),
                safe_format(row['best_x_source']),
                safe_format(row['best_2']),
                safe_format(row['best_2_source']),
                safe_format(row['implied']),
            ])
        detail_sheet = wb.create_sheet(title="Detailed Staking")
        detail_headers = [
            "Home", "Away", "Time",
            "Best_1", "Best_1_Source",
            "Best_X", "Best_X_Source",
            "Best_2", "Best_2_Source",
            "Stake1", "Stake2", "Stake3",
            "Profit1", "Profit2", "Profit3"
        ]
        detail_sheet.append(detail_headers)
        for _, row in sure_bets.iterrows():
            detail_sheet.append([
                safe_format(row['home']),
                safe_format(row['away']),
                safe_format(row['time']),
                safe_format(row['best_1']),
                safe_format(row['best_1_source']),
                safe_format(row['best_x']),
                safe_format(row['best_x_source']),
                safe_format(row['best_2']),
                safe_format(row['best_2_source']),
                safe_format(row['Stake1']),
                safe_format(row['Stake2']),
                safe_format(row['Stake3']),
                safe_format(row['Profit1']),
                safe_format(row['Profit2']),
                safe_format(row['Profit3']),
            ])
        output_path = os.path.join(config["OUTPUT_FOLDER"], "sure_bets_strict.xlsx")
        try:
            wb.save(output_path)
            logging.info(f"Sure bet results saved to {output_path}")
            print(f"Sure bet results saved to {output_path}")
        except Exception as e:
            logging.error(f"Failed to save Excel file. Error: {e}")
            print("Failed to save Excel file. Check logs for details.")
    else:
        logging.info("No sure bets found after implied probability calculation.")

    pkl_paths = glob.glob(os.path.join(config["PICKLE_FOLDER_PATH"], "*.pkl"))
    logging.info(f"Found {len(pkl_paths)} pickle files to archive.")
    for pkl in pkl_paths:
        archive_file(pkl, config["ARCHIVE_FOLDER"])

    try:
        excel_paths = [os.path.join(EXCEL_FOLDER_PATH, f)
                       for f in os.listdir(EXCEL_FOLDER_PATH)
                       if f.lower().endswith('.xlsx')]
        logging.info(f"Found {len(excel_paths)} Excel files to archive.")
        for xls in excel_paths:
            archive_file(xls, config["ARCHIVE_FOLDER"])
    except FileNotFoundError:
        logging.error(f"Excel folder '{EXCEL_FOLDER_PATH}' does not exist. No Excel files to archive.")
    except Exception as e:
        logging.error(f"Error while archiving Excel files: {e}")

if __name__ == "__main__":
    main()
